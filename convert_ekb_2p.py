"""One-time conversion script: convert ekb_2p Excel spectrometer files to standard CSV.

Run from the light-calibration directory:
    python convert_ekb_2p.py
"""

import os
import numpy as np
import openpyxl

from src.spectrum_utils import resample_spectrum, save_spectrum_csv, get_spectra_dir

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EKB_DIR = os.path.join(BASE_DIR, 'ekb_2p')


def read_excel_spectrum(filepath, sheet_name, wl_col, val_col, skip_rows=1):
    """Read wavelength and intensity columns from an Excel file.

    Parameters
    ----------
    filepath : str
        Path to the .xlsx file.
    sheet_name : str
        Name of the worksheet to read.
    wl_col : int
        1-based column index for wavelength data.
    val_col : int
        1-based column index for intensity data.
    skip_rows : int
        Number of header/metadata rows to skip (default 1).

    Returns
    -------
    wavelengths : ndarray
    intensities : ndarray
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    wavelengths = []
    intensities = []
    for row in ws.iter_rows(min_row=skip_rows + 1, max_row=ws.max_row,
                            min_col=wl_col, max_col=val_col, values_only=True):
        wl_val = row[0]
        int_val = row[val_col - wl_col]
        if wl_val is not None and int_val is not None:
            wavelengths.append(float(wl_val))
            intensities.append(float(int_val))

    wb.close()
    return np.array(wavelengths), np.array(intensities)


# Configuration for each Excel file
SPECTRA = [
    {
        'filename': '080724_AJ_UV_2 1.xlsx',
        'sheet': '080724_AJ_UV_2',
        'wl_col': 1,   # Column A
        'val_col': 2,   # Column B
        'output_name': 'lightCrafter_2p_uv',
    },
    {
        'filename': '080524_AJ_Green2.xlsx',
        'sheet': '080523_AJ_Green_F4',
        'wl_col': 4,   # Column D
        'val_col': 5,   # Column E
        'output_name': 'lightCrafter_2p_green',
    },
    {
        'filename': '103024_AJ_RED 1.xlsx',
        'sheet': 'RED Multiphoton F2',
        'wl_col': 1,   # Column A
        'val_col': 2,   # Column B
        'output_name': 'lightCrafter_2p_red',
    },
]


def main():
    output_dir = get_spectra_dir('stimuli')
    print("Converting ekb_2p Excel files to standard CSV:")

    for spec in SPECTRA:
        filepath = os.path.join(EKB_DIR, spec['filename'])
        if not os.path.exists(filepath):
            print(f"  SKIP (not found): {spec['filename']}")
            continue

        wavelengths, intensities = read_excel_spectrum(
            filepath, spec['sheet'], spec['wl_col'], spec['val_col']
        )

        print(f"  {spec['filename']}: {len(wavelengths)} points, "
              f"wl range {wavelengths[0]:.1f}-{wavelengths[-1]:.1f} nm")

        # Baseline correction: subtract minimum to floor at zero
        intensities = intensities - intensities.min()

        # Resample to standard grid (200-720nm, 1nm steps)
        new_wl, new_vals = resample_spectrum(wavelengths, intensities)

        # Normalize to sum to 1
        total = new_vals.sum()
        if total > 0:
            new_vals = new_vals / total

        outpath = os.path.join(output_dir, spec['output_name'] + '.csv')
        save_spectrum_csv(outpath, new_wl, new_vals,
                          header="wavelength_nm,relative_intensity")
        print(f"    -> {spec['output_name']}.csv")

    print("\nConversion complete.")


if __name__ == '__main__':
    main()
